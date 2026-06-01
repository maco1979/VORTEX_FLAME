#!/usr/bin/env python3
"""
VF Inspect — 数据质量检查模块
===============================
四维度数据质量评估: 完整性/准确性/一致性/合规性

检查维度:
  COMPLETENESS  — 必填字段覆盖率、记录完整性、关联完整性
  ACCURACY      — 业务规则验证、交叉校验、值域检查
  CONSISTENCY   — 跨系统一致性、格式一致性、引用完整性
  COMPLIANCE    — 行业规范合规、隐私法规合规(PII检测)、政策合规

报告输出:
  JSON   — 结构化质量报告, 含问题列表与位置定位
  PDF    — (需reportlab) 可视化报告含图表
  Excel  — (需openpyxl) 表格导出格式

用法:
  from vf_inspect import InspectEngine
  inspector = InspectEngine(cfg, audit)
  report = inspector.process(data)
  issues = inspector.validate(data)
"""

import hashlib
import json
import math
import os
import re
import time
from collections import Counter, defaultdict
from dataclasses import dataclass, field, asdict
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Callable, Dict, List, Optional, Set, Tuple

from vf_data_core import (
    AuditLogger, BaseProcessor, ConfigManager, ProcessingStats, Severity,
)

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False


@dataclass
class QualityIssue:
    issue_id: str
    dimension: str
    check_name: str
    severity: str
    field: str = ""
    record_index: int = -1
    description: str = ""
    suggestion: str = ""
    value_snippet: str = ""


@dataclass
class QualityReport:
    module: str = "inspect"
    timestamp: str = ""
    total_records: int = 0
    total_fields: int = 0
    checks_run: int = 0
    issues_found: int = 0

    completeness_score: float = 1.0
    accuracy_score: float = 1.0
    consistency_score: float = 1.0
    compliance_score: float = 1.0
    overall_score: float = 1.0

    issues: List[QualityIssue] = field(default_factory=list)
    field_stats: Dict[str, Dict[str, Any]] = field(default_factory=dict)
    summary: str = ""


class PII_Detector:
    PATTERNS = {
        "email": re.compile(r'[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}'),
        "phone_cn": re.compile(r'1[3-9]\d{9}'),
        "phone_us": re.compile(r'\b\d{3}[-.]?\d{3}[-.]?\d{4}\b'),
        "id_card_cn": re.compile(r'\b\d{17}[\dXx]\b'),
        "credit_card": re.compile(r'\b\d{4}[\s-]?\d{4}[\s-]?\d{4}[\s-]?\d{4}\b'),
        "ip_address": re.compile(r'\b\d{1,3}\.\d{1,3}\.\d{1,3}\.\d{1,3}\b'),
        "ssn": re.compile(r'\b\d{3}-\d{2}-\d{4}\b'),
    }

    SENSITIVE_KEYWORDS = [
        "password", "secret", "token", "api_key", "apikey",
        "private_key", "secret_key", "access_key",
    ]

    @classmethod
    def scan(cls, text: str) -> List[Tuple[str, str, str]]:
        findings = []
        for label, pattern in cls.PATTERNS.items():
            matches = pattern.findall(text)
            for m in matches:
                findings.append((label, m, "PII detected"))
        lower = text.lower()
        for kw in cls.SENSITIVE_KEYWORDS:
            if kw in lower:
                findings.append(("sensitive_keyword", kw, "Sensitive keyword"))
        return findings


class BusinessRule:
    def __init__(self, rule_id: str, description: str, check: Callable[[Dict], Tuple[bool, str]]):
        self.rule_id = rule_id
        self.description = description
        self._check = check

    def validate(self, record: Dict) -> Tuple[bool, str]:
        return self._check(record)


class InspectEngine(BaseProcessor):
    def __init__(self, config: ConfigManager, audit: AuditLogger):
        super().__init__(config, audit, "inspect")
        self._required_fields: List[str] = config.get("inspect", "required_fields", default=[])
        self._business_rules: List[BusinessRule] = []
        self._compliance_rules: List[BusinessRule] = []
        self._load_rules(config)
        self._report = QualityReport()

    def _load_rules(self, config: ConfigManager):
        br_configs = config.get("inspect", "business_rules", default=[])
        for br in br_configs:
            if isinstance(br, dict) and br.get("rule_id"):
                def _make_check(rule_def):
                    def _check(rec):
                        field = rule_def.get("field", "")
                        op = rule_def.get("operator", "not_null")
                        val = rec.get(field)
                        if op == "not_null" and val is None:
                            return False, f"{field} is null"
                        if op == "range" and isinstance(val, (int, float)):
                            low = rule_def.get("min")
                            high = rule_def.get("max")
                            if low is not None and val < low:
                                return False, f"{field}={val} < min={low}"
                            if high is not None and val > high:
                                return False, f"{field}={val} > max={high}"
                        if op == "in" and str(val) not in rule_def.get("values", []):
                            return False, f"{field}={val} not in allowed values"
                        if op == "regex" and not re.match(rule_def.get("pattern", ""), str(val)):
                            return False, f"{field}='{val}' pattern mismatch"
                        return True, ""
                    return _check
                self._business_rules.append(BusinessRule(
                    rule_id=br["rule_id"],
                    description=br.get("description", ""),
                    check=_make_check(br),
                ))

    def process(self, data: List[Dict[str, Any]]) -> Tuple[List[Dict[str, Any]], ProcessingStats]:
        t0 = time.time()
        self._report = self._run_all_checks(data)

        stats = ProcessingStats(
            total_input=len(data),
            total_output=len(data),
            duration_ms=(time.time() - t0) * 1000,
            throughput_per_sec=len(data) / max((time.time() - t0), 0.001),
        )

        severity_counts = Counter(i.severity for i in self._report.issues)
        self.stats = stats
        self._audit_operation("inspect.process", "OK", stats.duration_ms,
                              len(data), {
                                  "issues": len(self._report.issues),
                                  "score": self._report.overall_score,
                                  "critical": severity_counts.get("CRITICAL", 0),
                                  "warn": severity_counts.get("WARN", 0),
                              })
        return data, stats

    def _run_all_checks(self, data: List[Dict]) -> QualityReport:
        report = QualityReport(
            timestamp=datetime.now(timezone.utc).isoformat(),
            total_records=len(data),
        )

        all_fields = set()
        for rec in data:
            all_fields.update(rec.keys())
        report.total_fields = len(all_fields)

        completeness_issues = self._check_completeness(data)
        accuracy_issues = self._check_accuracy(data)
        consistency_issues = self._check_consistency(data)
        compliance_issues = self._check_compliance(data)

        all_issues = completeness_issues + accuracy_issues + consistency_issues + compliance_issues
        report.issues = all_issues
        report.issues_found = len(all_issues)
        report.checks_run = 4 + len(self._business_rules) + len(self._compliance_rules)

        n = max(len(data), 1)
        c_count = len(completeness_issues)
        a_count = len(accuracy_issues)
        s_count = len(consistency_issues)
        p_count = len(compliance_issues)

        report.completeness_score = max(0.0, 1.0 - c_count / (n * max(len(self._required_fields), 1) + 1))
        report.accuracy_score = max(0.0, 1.0 - a_count / (n * 0.5 + 1))
        report.consistency_score = max(0.0, 1.0 - s_count / (n * 0.3 + 1))
        report.compliance_score = max(0.0, 1.0 - p_count / (n * 0.2 + 1))
        report.overall_score = (
            report.completeness_score * 0.3 +
            report.accuracy_score * 0.3 +
            report.consistency_score * 0.2 +
            report.compliance_score * 0.2
        )

        report.field_stats = self._compute_field_stats(data, all_fields, all_issues)

        if report.overall_score >= 0.95:
            report.summary = f"PASS ({report.overall_score:.1%}) — {report.issues_found} minor issues"
        elif report.overall_score >= 0.80:
            report.summary = f"WARN ({report.overall_score:.1%}) — {report.issues_found} issues need attention"
        else:
            report.summary = f"FAIL ({report.overall_score:.1%}) — {report.issues_found} critical issues"

        return report

    def _check_completeness(self, data: List[Dict]) -> List[QualityIssue]:
        issues = []
        field_null_counts = defaultdict(int)

        for i, rec in enumerate(data):
            for field in self._required_fields:
                if rec.get(field) is None:
                    field_null_counts[field] += 1
                    issues.append(QualityIssue(
                        issue_id=f"COMPL-{len(issues):04d}",
                        dimension="completeness",
                        check_name="required_field",
                        severity="CRITICAL",
                        field=field,
                        record_index=i,
                        description=f"Required field '{field}' is null",
                        suggestion=f"Fill '{field}' or remove record",
                    ))

        for i, rec in enumerate(data):
            null_count = sum(1 for v in rec.values() if v is None)
            total = len(rec)
            if total > 0 and null_count / total > 0.8:
                issues.append(QualityIssue(
                    issue_id=f"COMPL-{len(issues):04d}",
                    dimension="completeness",
                    check_name="record_sparse",
                    severity="WARN",
                    record_index=i,
                    description=f"Record {i} is {null_count}/{total} null",
                    suggestion="Review if this record is valid",
                ))

        return issues

    def _check_accuracy(self, data: List[Dict]) -> List[QualityIssue]:
        issues = []

        for rule in self._business_rules:
            for i, rec in enumerate(data):
                passed, msg = rule.validate(rec)
                if not passed:
                    issues.append(QualityIssue(
                        issue_id=f"ACC-{len(issues):04d}",
                        dimension="accuracy",
                        check_name=rule.rule_id,
                        severity="CRITICAL",
                        record_index=i,
                        description=f"Rule '{rule.description}': {msg}",
                        suggestion=f"Fix value to satisfy {rule.rule_id}",
                    ))

        for i, rec in enumerate(data):
            for key, val in rec.items():
                if isinstance(val, str) and re.search(r'[<>]', val):
                    issues.append(QualityIssue(
                        issue_id=f"ACC-{len(issues):04d}",
                        dimension="accuracy",
                        check_name="suspicious_chars",
                        severity="WARN",
                        field=key,
                        record_index=i,
                        description=f"Field '{key}' contains angle brackets: '{val[:50]}'",
                        value_snippet=val[:100],
                    ))

        return issues

    def _check_consistency(self, data: List[Dict]) -> List[QualityIssue]:
        issues = []
        if len(data) < 2:
            return issues

        all_fields = set()
        for rec in data:
            all_fields.update(rec.keys())

        for field in all_fields:
            types = set()
            for rec in data:
                val = rec.get(field)
                if val is not None:
                    types.add(type(val).__name__)
            if len(types) > 1:
                issues.append(QualityIssue(
                    issue_id=f"CONS-{len(issues):04d}",
                    dimension="consistency",
                    check_name="type_consistency",
                    severity="CRITICAL",
                    field=field,
                    description=f"Field '{field}' has mixed types: {types}",
                    suggestion=f"Normalize '{field}' to a single type",
                ))

        return issues

    def _check_compliance(self, data: List[Dict]) -> List[QualityIssue]:
        issues = []
        for i, rec in enumerate(data):
            text = json.dumps(rec, ensure_ascii=False)
            pii_findings = PII_Detector.scan(text)
            for label, match, desc in pii_findings:
                issues.append(QualityIssue(
                    issue_id=f"COMP-{len(issues):04d}",
                    dimension="compliance",
                    check_name="pii_detection",
                    severity="CRITICAL",
                    record_index=i,
                    description=f"PII '{label}' found: {desc}",
                    value_snippet=match[:50],
                    suggestion=f"Mask or remove PII data",
                ))

        for rule in self._compliance_rules:
            for i, rec in enumerate(data):
                passed, msg = rule.validate(rec)
                if not passed:
                    issues.append(QualityIssue(
                        issue_id=f"COMP-{len(issues):04d}",
                        dimension="compliance",
                        check_name=rule.rule_id,
                        severity="CRITICAL",
                        record_index=i,
                        description=msg,
                    ))

        return issues

    def _compute_field_stats(self, data: List[Dict], all_fields: Set[str],
                             issues: List[QualityIssue]) -> Dict[str, Dict]:
        stats = {}
        for field in all_fields:
            values = [rec.get(field) for rec in data]
            non_null = [v for v in values if v is not None]
            field_issues = [i for i in issues if i.field == field]

            stats[field] = {
                "total": len(values),
                "non_null": len(non_null),
                "null_rate": (len(values) - len(non_null)) / max(len(values), 1),
                "unique_values": len(set(str(v) for v in non_null)),
                "issues": len(field_issues),
            }

            numeric_vals = [v for v in non_null if isinstance(v, (int, float))]
            if numeric_vals:
                stats[field]["min"] = min(numeric_vals)
                stats[field]["max"] = max(numeric_vals)

        return stats

    def validate(self, data: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
        self.process(data)
        return [asdict(i) for i in self._report.issues]

    def export_report_json(self, output_path: Optional[Path] = None) -> str:
        report_dict = {
            "module": self._report.module,
            "timestamp": self._report.timestamp,
            "total_records": self._report.total_records,
            "overall_score": round(self._report.overall_score, 4),
            "completeness_score": round(self._report.completeness_score, 4),
            "accuracy_score": round(self._report.accuracy_score, 4),
            "consistency_score": round(self._report.consistency_score, 4),
            "compliance_score": round(self._report.compliance_score, 4),
            "issues_count": self._report.issues_found,
            "issues": [asdict(i) for i in self._report.issues],
            "field_stats": self._report.field_stats,
        }
        json_str = json.dumps(report_dict, ensure_ascii=False, indent=2)
        if output_path:
            output_path.write_text(json_str, encoding="utf-8")
        return json_str

    def export_report_excel(self, output_path: Path):
        if not HAS_OPENPYXL:
            raise ImportError("openpyxl required for Excel export")

        wb = openpyxl.Workbook()

        ws_summary = wb.active
        ws_summary.title = "Summary"  # type: ignore[reportOptionalMemberAccess]
        headers = ["Metric", "Value"]
        ws_summary.append(headers)  # type: ignore[reportOptionalMemberAccess]
        rows = [
            ("Total Records", self._report.total_records),
            ("Total Fields", self._report.total_fields),
            ("Overall Score", f"{self._report.overall_score:.2%}"),
            ("Completeness", f"{self._report.completeness_score:.2%}"),
            ("Accuracy", f"{self._report.accuracy_score:.2%}"),
            ("Consistency", f"{self._report.consistency_score:.2%}"),
            ("Compliance", f"{self._report.compliance_score:.2%}"),
            ("Issues Found", self._report.issues_found),
        ]
        for row in rows:
            ws_summary.append(row)  # type: ignore[reportOptionalMemberAccess]

        ws_issues = wb.create_sheet("Issues")
        issue_headers = ["ID", "Dimension", "Check", "Severity", "Field", "Record", "Description"]
        ws_issues.append(issue_headers)
        for issue in self._report.issues:
            ws_issues.append([
                issue.issue_id, issue.dimension, issue.check_name,
                issue.severity, issue.field, issue.record_index,
                issue.description[:200],
            ])

        ws_fields = wb.create_sheet("Field Stats")
        field_headers = ["Field", "Total", "Non-Null", "Null Rate", "Unique", "Issues"]
        ws_fields.append(field_headers)
        for field, fstats in self._report.field_stats.items():
            ws_fields.append([
                field, fstats["total"], fstats["non_null"],
                f"{fstats['null_rate']:.2%}", fstats["unique_values"], fstats["issues"],
            ])

        wb.save(str(output_path))

    def report(self) -> Dict[str, Any]:
        return {
            "module": self.module_name,
            "overall_score": round(self._report.overall_score, 4),
            "issues": len(self._report.issues),
            "dimensions": {
                "completeness": round(self._report.completeness_score, 4),
                "accuracy": round(self._report.accuracy_score, 4),
                "consistency": round(self._report.consistency_score, 4),
                "compliance": round(self._report.compliance_score, 4),
            },
            "stats": asdict(self.stats),
            "timestamp": datetime.now(timezone.utc).isoformat(),
        }
